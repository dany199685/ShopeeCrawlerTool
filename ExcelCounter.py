import os
import copy
import globals as Global
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image

def GetItemImgSafely (file_list, itemImgName:str):
    bFind = bool ()
    name = itemImgName + ".jpg"
    for img_name in file_list:
        if img_name == name:
            bFind = True
            break
    if bFind:
        return Image (Global.DF_EXPORT_IMG_PATH + name)
    else:return None

def ExportOrder (file_list, currRow, shopId, order, sheet):
    # Layout setting
    # TODO:整張Excel都相同的設定可以改成統一處理 ex:欄位寬度
    sheet.column_dimensions[Global.get_value ("def_itemImg_col_index")].width    = Global.get_value ("def_itemImg_grid_width")
    sheet.column_dimensions[Global.get_value ("def_shopId_col_index")].width     = Global.get_value ("def_item_shopId_grid_width")
    sheet.column_dimensions[Global.get_value ("def_itemId_col_index")].width     = Global.get_value ("def_itemId_grid_width")
    sheet.column_dimensions[Global.get_value ("def_itemName_col_index")].width   = Global.get_value ("def_itemName_grid_width")
    sheet.column_dimensions[Global.get_value ("def_item_subId_col_index")].width  = Global.get_value ("def_itemSubId_grid_width")
    sheet.column_dimensions[Global.get_value ("def_itemNum_col_index")].width    = Global.get_value ("def_itemNum_grid_width")
    sheet.column_dimensions[Global.get_value ("def_itemDetail_col_index")].width = Global.get_value ("def_itemDetail_grid_width")
    
    AlignCenterType = Alignment(horizontal="center", vertical="center")
    export_list = list (sheet.rows)
    if len (export_list) == 0:
        sheet.append (Global.get_value ("def_excel_titles"))
        sheet[Global.get_value ("def_itemImg_col_index") + "1"].alignment = AlignCenterType
        sheet[Global.get_value ("def_shopId_col_index") + "1"].alignment = AlignCenterType
        sheet[Global.get_value ("def_itemId_col_index") + "1"].alignment = AlignCenterType
        sheet[Global.get_value ("def_itemName_col_index") + "1"].alignment = AlignCenterType
        sheet[Global.get_value ("def_item_subId_col_index") + "1"].alignment = AlignCenterType
        sheet[Global.get_value ("def_itemNum_col_index") + "1"].alignment = AlignCenterType
        sheet[Global.get_value ("def_itemDetail_col_index") + "1"].alignment = AlignCenterType


    order_size = len (order.m_SubItemList)
    BeginRow = currRow + 1
    EndRow = currRow + order_size

    # 儲存格合併
    strBeginRow = str (BeginRow)
    strEndRow = str (EndRow)
    sheet.merge_cells (Global.get_value ("def_shopId_col_index") + strBeginRow + ":" + Global.get_value ("def_shopId_col_index") + strEndRow)     # 共同欄位 - 賣場編號
    sheet.merge_cells (Global.get_value ("def_itemId_col_index") + strBeginRow + ":" + Global.get_value ("def_itemId_col_index") + strEndRow)     # 共同欄位 - 商品主要貨號
    sheet.merge_cells (Global.get_value ("def_itemName_col_index") + strBeginRow + ":" + Global.get_value ("def_itemName_col_index") + strEndRow) # 共同欄位 - 商品名稱

    # 對齊
    sheet[Global.get_value ("def_shopId_col_index") + strBeginRow].alignment = AlignCenterType
    sheet[Global.get_value ("def_itemId_col_index") + strBeginRow].alignment = AlignCenterType
    sheet[Global.get_value ("def_itemName_col_index") + strBeginRow].alignment = AlignCenterType

    # 寫入資料
    AlignLeftType = Alignment(horizontal="left", vertical="center")
    sheet[Global.get_value ("def_shopId_col_index") + strBeginRow] = shopId
    sheet[Global.get_value ("def_itemId_col_index") + strBeginRow] = order.m_ItemMainId
    sheet[Global.get_value ("def_itemName_col_index") + strBeginRow] = order.m_ItemName
    tempRow = BeginRow
    for subId in order.m_SubItemList:
        strRow = str (tempRow)

        sheet[Global.get_value ("def_item_subId_col_index") + strRow].alignment = AlignCenterType
        sheet[Global.get_value ("def_item_subId_col_index") + strRow] = subId # 商品選項貨號

        sheet[Global.get_value ("def_itemNum_col_index") + strRow].alignment = AlignCenterType
        sheet[Global.get_value ("def_itemNum_col_index") + strRow] = order.m_SubItemList[subId].m_ItemSubNum # 商品數量

        # 訂單詳細資料放同一儲存格，並且用特殊符號分開顯示
        strDetail = ""
        for s in order.m_SubItemList[subId].m_DetailInfos:
            if strDetail != "":
                strDetail += (Global.get_value ("def_itemDetail_spliter") + s)
            else:
                strDetail = s
        
        sheet[Global.get_value ("def_itemDetail_col_index") + strRow].alignment = AlignLeftType
        sheet[Global.get_value ("def_itemDetail_col_index") + strRow] = strDetail # 商品選項貨號
        tempRow += 1

    # 等合併好商品列數後再設定
    # 合併欄位(縱向)儲存格時，格子座標為變成所有合併的儲存格的第一格 Ex:[A1:A5]合併 > 合併後的儲存格為[A1]
    if order_size < Global.get_value ("def_grid_autoSetHeight_by_minOrderSize"):
        # 未停整過的列高為40
        # 調整其他子列高成平均高度
        subRowSize = EndRow - BeginRow + 1
        for i in range (BeginRow, EndRow + 1):
            sheet.row_dimensions[i].height = Global.get_value ("def_itemImg_grid_min_height") / subRowSize

    # 圖片會在儲存格左上角生成
    sheet.merge_cells (Global.get_value ("def_itemImg_col_index") + strBeginRow + ":" + Global.get_value ("def_itemImg_col_index") + strEndRow)   # 共同欄位 - 商品參考圖
    img = GetItemImgSafely (file_list, order.m_ItemName)
    if img != None:
        img.width = Global.get_value ("def_itemImg_width")
        img.height = Global.get_value ("def_itemImg_height")
        sheet.add_image (img, Global.get_value ("def_itemImg_col_index") + strBeginRow)
    return EndRow

def GetExcelFromAccount (file_list, account, sheet, nCurrRow = 1):
    nOldRow = nCurrRow
    for shopId in account.m_Shops:
        Shop = account.m_Shops[shopId]
        for itemId in Shop.m_Orders:
            order = Shop.m_Orders[itemId]
            nCurrRow = ExportOrder (file_list, nCurrRow, shopId, order, sheet)

    dirty = nOldRow != nCurrRow    
    # 增加篩選器
    # Fix:目前在使用篩選功能時，因為列在合併時，會讓某些儲存格的值為空值，導致篩選資料錯誤
    # Fix:圖片無法經由篩選器正確忽略
    if dirty:
        order_list = list (sheet.rows)
        strOrderSize = str (len (order_list))
        sheet.auto_filter.ref = Global.get_value ("def_shopId_col_index")+ "1:" + Global.get_value ("def_shopId_col_index") + strOrderSize
    return nCurrRow

def CombineAccounts (Accounts):
    newAccount = CShopeeAccount ()
    newAccount.m_AccountId = "" # 全部統計暫不分類賣場Id
    for accountId in Accounts:
        Account = Accounts[accountId]
        for shopId in Account.m_Shops:
            shop = Account.m_Shops[shopId]
            for itemId in shop.m_Orders:
                order = shop.m_Orders[itemId]
                if shopId in newAccount.m_Shops:
                    Shop = newAccount.m_Shops[shopId]
                    if itemId in Shop.m_Orders:
                        Order = Shop.m_Orders[itemId]
                        for subId in order.m_SubItemList:
                            orderdetail = order.m_SubItemList[subId]
                            if subId in Order.m_SubItemList:
                                OrderDetail = Order.m_SubItemList[subId]
                                OrderDetail.m_ItemSubNum += orderdetail.m_ItemSubNum
                                for v in orderdetail.m_DetailInfos:
                                    OrderDetail.m_DetailInfos.append (v)
                            else:
                                OrderDetail = COrderDetailInfo ()
                                OrderDetail = copy.deepcopy (orderdetail)
                                Order.m_SubItemList[subId] = OrderDetail
                    else:
                        Order = COrderInfo ()
                        Order = copy.deepcopy (order)
                        Shop.m_Orders[itemId] = Order
                else:
                    Shop = CShop (shopId)
                    Order = COrderInfo ()
                    Order = copy.deepcopy (order)
                    Shop.m_Orders[itemId] = Order
                    newAccount.m_Shops[shopId] = Shop

        
        for v in Account.m_ErrOrderList:
            newAccount.m_ErrOrderList.append (v)
    
    return newAccount

def GetParamsString (*args):
    strResult = ""
    for v in args:
        if strResult == "":
            strResult = str (v)
        else:
            strResult += "_"
            strResult += str (v)
    return strResult



class COrderDetailInfo ():
    m_ItemSubId     = None
    m_ItemSubName   = None
    m_ItemSubNum    = None
    m_DetailInfos   = None # <訂單詳細資訊:賣家後台Id(Excel名稱)_買家帳號_訂單編號_數量>

    def __init__ (self, _itemSubId = str ()):
        self.m_ItemSubId     = _itemSubId
        self.m_ItemSubName   = str ()
        self.m_ItemSubNum    = int ()
        self.m_DetailInfos   = list ()

class COrderInfo ():
    m_ItemMainId    = None
    m_ItemName      = None
    m_SubItemList   = None # <ItemOptionID, COrderDetailInfo> EX:ItemOptionID = "black_m"

    def __init__ (self, _itemMainId = str ()):
        self.m_ItemMainId    = _itemMainId
        self.m_ItemName      = str ()
        self.m_SubItemList   = dict ()

class CShop ():
    m_ShopId    = None
    m_ShopName  = None
    m_Orders    = None # <ItemMainID, COrderInfo>  EX:ItemMainID = "10001"

    def __init__ (self, _shopId = str ()):
        self.m_ShopId    = _shopId
        self.m_ShopName  = str ()
        self.m_Orders    = dict ()
    
class CShopeeAccount ():
    m_AccountId     = None
    m_WorkBook      = None
    m_WorkSheet     = None
    m_Shops         = None # <"ShopID, CShop"> EX:ShopID = "A"
    m_ErrOrderList  = None

    def __init__ (self, *args):
        if type (args) != tuple:
            return
        
        self.m_Shops = dict ()
        self.m_ErrOrderList = list ()
        if len (args) == 2:
            if type (args[1]) != Workbook:
                print ("Lack the requested workbook!")
                return
            self.m_AccountId = args[0]
            self.m_WorkBook = args[1]
            if Global.get_value ("def_worksheet_default_name") in self.m_WorkBook.sheetnames:
                self.m_WorkSheet = self.m_WorkBook.get_sheet_by_name (Global.get_value ("def_worksheet_default_name"))
            

    def DoWorkSheetProcess (self):
        if self.m_WorkSheet == None:
            return
        
        order_list = list (self.m_WorkSheet.rows)
        if len (order_list) <= 1:
            return

        # Classify items by item_main_id and item_option_id.
        for i in range (1, len (order_list)): # 0是title
            # 主要商品貨號:廠商_主要商品貨號
            # 商品選項貨號:商品選項1_商品選項2 (不須拆開，僅供識別)
            orderInfo = [item.value for item in order_list[i]]

            strItemMainId = orderInfo[Global.get_value ("def_itemMainId_col")]
            if type (strItemMainId) != str:
                continue

            strSplit = strItemMainId.split ("_")
            if len (strSplit) != Global.get_value ("def_itemMainParams_num"):
                self.m_ErrOrderList.append (orderInfo)
                continue
            
            shopId = strSplit[0]
            itemId = strSplit[1]
            orderId = orderInfo[Global.get_value ("def_itemId_col")]
            buyerId = orderInfo[Global.get_value ("def_buyerId_col")]
            itemName = orderInfo[Global.get_value ("def_itemName_col")]
            ItemOptionId = orderInfo[Global.get_value ("def_itemOptionId_col")]
            item_num = orderInfo[Global.get_value ("def_itemNum_col")]
            strDetail = GetParamsString (self.m_AccountId, buyerId, orderId, item_num)

            if self.m_Shops.get (shopId, None) == None:
                shop = CShop (shopId)
                order = COrderInfo (itemId)
                order.m_ItemName = itemName
                detailInfo = COrderDetailInfo (ItemOptionId)
                detailInfo.m_ItemSubNum = item_num
                detailInfo.m_DetailInfos.append (strDetail)

                order.m_SubItemList[ItemOptionId] = detailInfo
                shop.m_Orders[itemId] = order
                self.m_Shops [shopId] = shop
            else:
                shop = self.m_Shops[shopId]
                if shop.m_Orders.get (itemId, None) == None:
                    order = COrderInfo (itemId)
                    order.m_ItemName = itemName
                    detailInfo = COrderDetailInfo (ItemOptionId)
                    detailInfo.m_ItemSubNum = item_num
                    detailInfo.m_DetailInfos.append (strDetail)

                    order.m_SubItemList[ItemOptionId] = detailInfo
                    shop.m_Orders[itemId] = order
                else:
                    order = shop.m_Orders[itemId]
                    if order.m_SubItemList.get (ItemOptionId, None) == None:
                        detailInfo = COrderDetailInfo (ItemOptionId)
                        detailInfo.m_ItemSubNum = item_num
                        detailInfo.m_DetailInfos.append (strDetail)
                        order.m_SubItemList[ItemOptionId] = detailInfo
                    else:
                        detailInfo = order.m_SubItemList[ItemOptionId]
                        detailInfo.m_ItemSubNum += item_num
                        detailInfo.m_DetailInfos.append (strDetail)

def ReadAllShopeeOrderExcel (super, _AllShopeeAccounts):
    # MainAction
    # 讀取來源路徑下的Excel，<key:accountId, value:excel>
    file_list = os.listdir (Global.DF_SOURCE_EXCEL_PATH)
    excel_list = dict ()
    for file in file_list:
        strArray = file.split (".")
        if strArray[1] == "xlsx":
            excel_list[strArray[0]] = load_workbook (Global.DF_SOURCE_EXCEL_PATH + file)

    for accountId in excel_list:
        shopeeAccount = CShopeeAccount (accountId, excel_list[accountId])
        shopeeAccount.DoWorkSheetProcess ()
        _AllShopeeAccounts [accountId] = shopeeAccount
    super.PrintLog ("Finish loading all excels.")

def ExportAllExcelData (super, _AllShopeeAccounts):
    # index_begin:全部帳號統計
    # index_n:各帳戶統計
    # index_end:錯誤訂單統計
    wb = Workbook ()
    sheet_index = 0
    file_list = os.listdir (Global.DF_EXPORT_IMG_PATH)

    # 所有帳號賣場統計
    TotalAccount = CombineAccounts (_AllShopeeAccounts)
    strFirstSheet = "Sheet"
    sheet = None
    if strFirstSheet in wb:
        sheet = wb[strFirstSheet]
    else:
        sheet = wb.create_sheet (title = strFirstSheet, index = sheet_index)
    sheet.title = "所有帳戶訂單統計"
    GetExcelFromAccount (file_list, TotalAccount, sheet)

    # 各別帳號賣場統計
    sheet_index += 1
    for accountId in _AllShopeeAccounts:
        sheet = wb.create_sheet (title = accountId, index = sheet_index)
        Account = _AllShopeeAccounts[accountId]
        GetExcelFromAccount (file_list, Account, sheet)
        sheet_index += 1

    # 錯誤格式訂單統計TotalAccount.m_ErrOrderList
    if len (TotalAccount.m_ErrOrderList) > 0:
        sheet = wb.create_sheet (title = "錯誤訂單彙整", index = sheet_index)

        # 塞標題
        for accountId in _AllShopeeAccounts:
            Account = _AllShopeeAccounts[accountId]
            order_list = list (Account.m_WorkSheet.rows)
            orderInfo = [item.value for item in order_list[0]]
            sheet.append (orderInfo)
            break
        
        for order in TotalAccount.m_ErrOrderList:
            sheet.append (order)
        
        # TODO:可考慮修飾文字排版 (儲存格寬度自適應)

    wb.save (Global.DF_EXPORT_EXCEL_PATH + Global.get_value ("def_export_excel_name") + ".xlsx")
    super.PrintLog ("Finish excel orders processing.")